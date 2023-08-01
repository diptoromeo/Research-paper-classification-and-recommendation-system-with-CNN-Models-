import matplotlib
import nltk
import openpyxl
import os
import regex
import pandas as pd

os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

from gensim.models import Word2Vec
from math import log
from nltk.tokenize import word_tokenize

from mpl_toolkits.mplot3d import axes3d
import matplotlib.pyplot as plt
from scipy.interpolate import interp2d
from matplotlib.ticker import LinearLocator, FormatStrFormatter
from matplotlib import cm
from matplotlib import pyplot
import matplotlib as mpl

#from networkx.drawing.tests.test_pylab import plt
from wordcloud import WordCloud

matplotlib.use('tkagg')

from nltk import Counter
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
from string import punctuation
import numpy as np

from numpy import array

from tensorflow.keras.preprocessing.text import Tokenizer
from tensorflow.keras.preprocessing.sequence import pad_sequences

from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Dense, Dropout
from tensorflow.keras.layers import Flatten
from tensorflow.keras.layers import Embedding
from tensorflow.keras.layers import Conv1D
from tensorflow.keras.layers import MaxPooling1D
from tensorflow.keras.layers import GlobalMaxPool1D
from tensorflow.keras.layers import Input
from tensorflow.keras.layers import concatenate
from tensorflow.keras.models import Model
from tensorflow.keras.utils import plot_model
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.metrics import precision_score
from sklearn.metrics import recall_score
from sklearn.metrics import f1_score
from sklearn.metrics import precision_recall_curve, average_precision_score
from sklearn.metrics import auc
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import roc_curve
from sklearn.metrics import roc_auc_score
from sklearn.feature_extraction.text import TfidfVectorizer
#
#
#
# path = "C:\Selenium practice\papers_1.xlsx"
#
# workbook = openpyxl.load_workbook(path)
#
# sheet = workbook.active
#
# rows = sheet.max_row
# cols = sheet.max_column
#
#
#
# print('\n----모든 행과 열 출력----')
# all_values = []
# for rows in sheet.rows:
#     rows_value = []
#     for cell in rows:
#         rows_value.append(cell.value)
#     print(rows_value)
#     all_values.append(rows_value)
#
# ##text write...................................................................................
# file1 = open("FGFSJournal.txt", "w", encoding='UTF8')
# items = all_values[0]
# for rows in all_values[1:]:
#     rows_value = []
#     count = 0
#     for cell in rows:
#         if count == 0:
#             file1.write("<paper " + items[count] + "=" + str(cell) + ">\n")
#         elif count == 1:
#             file1.write("\t<journal_theme>"+cell+"</journal_theme>\n")
#         elif count == 2:
#             file1.write("\t<paper_heading>"+cell+"</paper_heading>\n")
#         elif count == 3:
#             file1.write("\t<authors>"+cell+"</authors>\n")
#         elif count == 4:
#             file1.write("\t<abstract>"+cell+"</abstract>\n")
#         elif count == 5:
#             file1.write("\t<keywords>"+cell+"</keywords>\n")
#         elif count == 6:
#             file1.write("\t<publication_month_year>"+ str(cell)+"</publication_month_year>\n")
#         elif count == 7:
#             file1.write("\t<volumes_issues>" + cell + "</volumes_issues>\n")
#             file1.write("</paper>\n")
#         count += 1
#

# # #
# # # ###Test data from FGFS.................................................
# # # path = "C:\Selenium practice\FGFS.xlsx"
# # #
# # # workbook = openpyxl.load_workbook(path)
# # #
# # # sheet = workbook.active
# # #
# # # rows = sheet.max_row
# # # cols = sheet.max_column
# # #
# # #
# # #
# # # print('\n----모든 행과 열 출력----')
# # # test1_values = []
# # # for rows in sheet.rows:
# # #     rows_value = []
# # #     for cell in rows:
# # #         rows_value.append(cell.value)
# # #     print(rows_value)
# # #     test1_values.append(rows_value)
# # #
# # # ##text write...................................................................................
# # # file1 = open("TestData_FGFS.txt", "w", encoding='UTF8')
# # # items = test1_values[0]
# # # for rows in test1_values[1:]:
# # #     rows_value = []
# # #     count = 0
# # #     for cell in rows:
# # #         if count == 0:
# # #             file1.write("<paper " + items[count] + "=" + str(cell) + ">\n")
# # #         elif count == 1:
# # #             file1.write("\t<journal_theme>"+cell+"</journal_theme>\n")
# # #         elif count == 2:
# # #            file1.write("\t<paper_heading>"+cell+"</paper_heading>\n")
# # #         elif count == 3:
# # #            file1.write("\t<author>"+cell+"</author>\n")
# # #         elif count == 4:
# # #            file1.write("\t<abstract>"+cell+"</abstract>\n")
# # #         elif count == 5:
# # #           file1.write("\t<keyword>"+cell+"</keyword>\n")
# # #         elif count == 6:
# # #           file1.write("\t<publication_year>"+cell+"</publication_year>\n")
# # #           file1.write("</paper>\n")
# # #         count += 1
# # #
# # #
# # #
# # # path = "C:\Selenium practice\AWR.xlsx"
# # #
# # # workbook = openpyxl.load_workbook(path)
# # #
# # # sheet = workbook.active
# # #
# # # rows = sheet.max_row
# # # cols = sheet.max_column
# # #
# # #
# # #
# # # print('\n----모든 행과 열 출력----')
# # # test2_values = []
# # # for rows in sheet.rows:
# # #     rows_value = []
# # #     for cell in rows:
# # #         rows_value.append(cell.value)
# # #     print(rows_value)
# # #     test2_values.append(rows_value)
# # #
# # # #text write...................................................................................
# # # file1 = open("TestData_AWR.txt", "w", encoding='UTF8')
# # # items = test2_values[0]
# # # for rows in test2_values[1:]:
# # #     rows_value = []
# # #     count = 0
# # #     for cell in rows:
# # #         if count == 0:
# # #             file1.write("<paper " + items[count] + "=" + str(cell) + ">\n")
# # #         elif count == 1:
# # #             file1.write("\t<journal_theme>"+cell+"</journal_theme>\n")
# # #         elif count == 2:
# # #            file1.write("\t<paper_heading>"+cell+"</paper_heading>\n")
# # #         elif count == 3:
# # #            file1.write("\t<author>"+cell+"</author>\n")
# # #         elif count == 4:
# # #            file1.write("\t<abstract>"+cell+"</abstract>\n")
# # #         elif count == 5:
# # #           file1.write("\t<keyword>"+cell+"</keyword>\n")
# # #         elif count == 6:
# # #           file1.write("\t<publication_year>"+cell+"</publication_year>\n")
# # #           file1.write("</paper>\n")
# # #         count += 1
# # #
# # #
# ###Nltk abstract_words tokenize...........................................................................
# with open('FGFSJournal.txt', 'rt', encoding='UTF8') as file:
#     all_abstract = []
#     for line in file:
#         if '<abstract>' in line:
#             abstract = line.split('</abstract>')[0].split('<abstract>')[-1]
#             abstract = ''.join(i for i in abstract if not i.isdigit())
#             abstract = regex.sub('[^\w\d\s]+', '', abstract)
#             ##abstract = nltk.sent_tokenize(abstract)
#             abstract = nltk.word_tokenize(abstract)
#             stop_words = set(stopwords.words('english'))
#             filtered_sentence_abstract = [w.lower() for w in abstract if
#                                           w.lower() not in punctuation and w.lower() not in stop_words]
#             tagged_list = nltk.pos_tag(filtered_sentence_abstract)
#             nouns_list = [t[0] for t in tagged_list if t[-1] == 'NN']
#             lm = WordNetLemmatizer()
#             singluar_form = [lm.lemmatize(w, pos='v') for w in nouns_list]
#             all_abstract.append(singluar_form)
#
#print(all_abstract)
#
# ## CBOW.......................................................................................
# model = Word2Vec(sentences=all_abstract, size=100, window=5, min_count=50, workers=4, iter=100, sg=0)
# model.wv.save_word2vec_format('100-dimension.txt', binary=False)
#
#
# maximum = len(model.wv.vocab)
# all_abstract_Words = []
# for i in range(0, maximum):
#     all_abstract_Words.append(model.wv.index2word[i])
#
# all_abstract_Words.remove('paper')
# all_abstract_Words.remove('article')
# all_abstract_Words.remove('approach')
# all_abstract_Words.remove('performance')





## ##CNNs training........................................................................................
with open('ReadTexT.txt', 'rt', encoding='UTF8') as file:
    #total_sentences = []
    #total_words = []
    all_abstract = []
    for line in file:
        if '<abstract>' in line:
            abstract = line.split('</abstract>')[0].split('<abstract>')[-1]
            #total_words.extend(abstract)
            # abstract = nltk.sent_tokenize(abstract)
            # total_sentences.extend(abstract)
            abstract = ''.join(i for i in abstract if not i.isdigit())
            abstract = regex.sub('[^\w\d\s]+', '', abstract)

            abstract = nltk.word_tokenize(abstract)

            stop_words = set(stopwords.words('english'))
            filtered_sentence_abstract = [w.lower() for w in abstract if
                                          w.lower() not in punctuation and w.lower() not in stop_words]
            tagged_list = nltk.pos_tag(filtered_sentence_abstract)
            nouns_list = [t[0] for t in tagged_list if t[-1] == 'NN']
            lm = WordNetLemmatizer()
            singluar_form = [lm.lemmatize(w, pos='v') for w in nouns_list]
            all_abstract.extend(singluar_form)

#print("total sentences", len(total_sentences))
#print("total words", len(total_words))
#print("prepair words", len(all_abstract))
#
#
# #
# # ## CBOW.......................................................................................
# model = Word2Vec(sentences=all_abstract, size=100, window=5, min_count=1, workers=4, iter=100, sg=0)
# model.wv.save_word2vec_format('100-dimension.txt', binary=False)
#
#
# maximum = len(model.wv.vocab)
# all_abstract_Words = []
# for i in range(0, maximum):
#     all_abstract_Words.append(model.wv.index2word[i])


#print(len(all_abstract_Words))
# #pre-traind word2vec data.......................................................
# CBOW_embeddings = Word2Vec(sentences=all_abstract, size=100, window=5, workers=4, iter=100, sg=0)#
# #CBOW_embeddings = Word2Vec(sentences=all_abstract, size=100, min_count=1)
# CBOW_embeddings.wv.save_word2vec_format('CBOW_Pre-trained_word2Vec.txt', binary=False)
#
# maximum = len(CBOW_embeddings.wv.vocab)
# all_abstract_Words = []
# for i in range(0, maximum):
#     all_abstract_Words.append(CBOW_embeddings.wv.index2word[i])
#
#print('all_abstract_Words:', len(all_abstract_Words))
#
#

##Prepairinng FGFS test data...........................................................................
with open('TestData_FGFS.txt', 'rt', encoding='UTF8') as file:
    test_FGFS = []
    for line in file:
        if '<abstract>' in line:
            abstract = line.split('</abstract>')[0].split('<abstract>')[-1]
            abstract = ''.join(i for i in abstract if not i.isdigit())
            abstract = regex.sub('[^\w\d\s]+', '', abstract)
            abstract = nltk.word_tokenize(abstract)
            stop_words = set(stopwords.words('english'))
            filtered_sentence_abstract = [w.lower() for w in abstract if
                                          w.lower() not in punctuation and w.lower() not in stop_words]
            tagged_list = nltk.pos_tag(filtered_sentence_abstract)
            nouns_list = [t[0] for t in tagged_list if t[1] == 'NN']
            lm = WordNetLemmatizer()
            singluar_form = [lm.lemmatize(w, pos='v') for w in nouns_list]
            test_FGFS.append(singluar_form)


#print(max(map(len, test_FGFS)))
#print("Test data from FGFS:", test_FGFS)

#

##FGFS_labels..................................................................
select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
FGFS_labels = []
for i in range(0, 100):
    count = 0
    for j in range(0, len(select_words)):
        if select_words[j] in test_FGFS[i]:
            count += 1
    if count >= 2:
        FGFS_labels.append(1)
    else:
        FGFS_labels.append(0)


#print("FGFS_labels:", FGFS_labels)

print()
token = Tokenizer()  # create the tokenizer
token.fit_on_texts(test_FGFS)  # fit the tokenizer on the documents

word_index = token.word_index
#print('unique words: {}'.format(len(word_index)))


max_length = 259
test = token.texts_to_sequences(test_FGFS)
x_test = pad_sequences(test, maxlen=max_length, padding='post')
y_test = np.array(FGFS_labels)

#
#
#
# # # # #Prepairinng AWR test data...........................................................................
# with open('TestData_AWR.txt', 'rt', encoding='UTF8') as file:
#     test_AWR = []
#     for line in file:
#         if '<abstract>' in line:
#             abstract = line.split('</abstract>')[0].split('<abstract>')[-1]
#             abstract = ''.join(i for i in abstract if not i.isdigit())
#             abstract = regex.sub('https?:\/\/.*[\r\n]*', '', abstract) ## URL 제거
#             abstract = regex.sub('[^\w\d\s]+', '', abstract)
#             abstract = nltk.word_tokenize(abstract)
#             stop_words = set(stopwords.words('english'))
#             filtered_sentence_abstract = [w.lower() for w in abstract if
#                                           w.lower() not in punctuation and w.lower() not in stop_words]
#             tagged_list = nltk.pos_tag(filtered_sentence_abstract)
#             nouns_list = [t[0] for t in tagged_list if t[-1] == 'NN']
#             lm = WordNetLemmatizer()
#             singluar_form = [lm.lemmatize(w, pos='v') for w in nouns_list]
#             test_AWR.append(singluar_form)
#

#print(max(map(len, test_AWR)))
#print()
#print("Test data from AWR:", test_AWR)

#
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# AWR_labels = []
# for i in range(0, 100):
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in test_AWR[i]:
#             count += 1
#     if count >= 2:
#         AWR_labels.append(1)
#     else:
#         AWR_labels.append(0)
# print("AWR_labels:", AWR_labels)
#
# print()
# token = Tokenizer()  # create the tokenizer
# token.fit_on_texts(test_AWR)  # fit the tokenizer on the documents
#
# word_index = token.word_index
# #print('unique words: {}'.format(len(word_index)))
#
#
# max_length = 259
# test1 = token.texts_to_sequences(test_AWR)
# x_test_1 = pad_sequences(test1, maxlen=max_length, padding='post')
# y_test_1 = np.array(AWR_labels)
# print("AWR test label:", y_test_1)

#
# ##abstract CNN training...................................................
print("create the tokenizer")
token = Tokenizer()  # create the tokenizer
token.fit_on_texts(all_abstract)  # fit the tokenizer on the documents
#print(token.word_index)


word_index = token.word_index
#print('unique words: {}'.format(len(word_index)))

# # print()
vocab_size = len(token.word_index) + 1  # define vocabulary size (largest integer value)
#print('Vocabulary size: %d' % vocab_size)


max_length = 259
train, valid = train_test_split(all_abstract, test_size=0.30, random_state=1)

print("train", len(train))
print("valid", len(valid))

##train_labels.......................................................................................................
select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
train_labels = []
for i in range(0, 700):
    count = 0
    for j in range(0, len(select_words)):
        if select_words[j] in all_abstract[i]:
            count += 1
    if count >= 1:
        train_labels.append(1)
    else:
        train_labels.append(0)


### validation labels.........................................
select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
valid_labels = []
for i in range(0, 300):

    count = 0
    for j in range(0, len(select_words)):
        if select_words[j] in valid[i]:
            count += 1
    if count >= 1:
        valid_labels.append(1)
    else:
        valid_labels.append(0)

#print(valid_labels)


train_data = token.texts_to_sequences(train)
valid_data = token.texts_to_sequences(valid)
# print("integer incode:", data)
# print("length: ", len(data))

X_train = pad_sequences(train_data, maxlen=max_length, padding='post')
y_train = np.array(train_labels)
#print(len(X_train))

X_valid = pad_sequences(valid_data, max_length, padding='post')
y_valid = np.array(valid_labels)



# ## CNNs model with word2vec ....................................................................
embedding_index = {}
list_v = []
file = open('CBOW_Pre-trained_word2Vec.txt', 'rt', encoding='UTF8')
line = file.readline()
totalWords, numOfFeatures = line.split()
print(totalWords, numOfFeatures)
for line in file:
    values = line.split()
    list_v.append(values)
    word = values[0]
    coefs = array(values[1:], dtype='float64')
    embedding_index[word] = coefs
#file.close()

print('Found %s word vectors.' % len(embedding_index))
df_values = pd.DataFrame(list_v)
print(df_values, "\n")

# unique_words = len(word_index)
# total_words = unique_words + 1
# embedding_dim = 100

embedding_matrix1 = np.array([[0 for col in range(100)] for row in range(4787)])

for word, i in token.word_index.items():
    # try:
    embedding_vector = embedding_index.get(word)
    if embedding_vector is not None:
        if( i == 100):
            print(i,"번째 완료")
        for j in range(0, 100):
           embedding_matrix1[i][j] = embedding_vector[j]
        #print(i,"번째 완료")

# print("embedding_matrix1:", embedding_matrix1)


# #CBOW_CNNs word2vec model ................................................................................
embedding_dim = 100

model = Sequential()
inputs1 = Input(shape=(max_length,))
embedding1 = Embedding(vocab_size, 100, weights=[embedding_matrix1], input_length=max_length, trainable=False)(inputs1)
conv1 = Conv1D(filters=32, kernel_size=2, activation='relu')(embedding1)
drop1 = Dropout(0.3)(conv1)
#pool1 = MaxPooling1D(pool_size=2)(drop1)
pool1 = GlobalMaxPool1D()(drop1)
flat1 = Flatten()(pool1)
#channel 2
conv2 = Conv1D(filters=32, kernel_size=3, activation='relu')(embedding1)
drop2 = Dropout(0.3)(conv2)
#pool2 = MaxPooling1D(pool_size=2)(drop2)
pool2 = GlobalMaxPool1D()(drop2)
flat2 = Flatten()(pool2)
#channel 3
conv3 = Conv1D(filters=32, kernel_size=4, activation='relu')(embedding1)
drop3 = Dropout(0.3)(conv3)
#pool3 = MaxPooling1D(pool_size=2)(drop3)
pool3 = GlobalMaxPool1D()(drop3)
flat3 = Flatten()(pool3)
#merge
merged = concatenate([flat1, flat2, flat3])
#merged = concatenate([pool1, pool2, pool3])
#interpretation
dense1 = Dense(10, activation='relu')(merged)
outputs = Dense(1, activation='sigmoid')(dense1)

model = Model(inputs=[inputs1], outputs=outputs)
# compile
model.compile(loss='binary_crossentropy', optimizer='adam', metrics=['acc'])
# summarize
print(model.summary())
plot_model(model, to_file='Word2Vec_Multichannel.png', show_shapes=True, show_layer_names=True)
pyplot.show()


CBOW_multi = model.fit(X_train, y_train,  validation_data=(X_valid, y_valid), epochs=100, verbose=1)
CBOW_score = model.evaluate(X_train, y_train,  verbose=0)
print(('Accuracy: %f' % (CBOW_score[1]*100)))
CBOW_score1 = model.evaluate(x_test, y_test, verbose=0)
print(('FGFS Test Accuracy: %f' % (CBOW_score1[1]*100)))
#CBOW_score2 = model.evaluate(x_test_1, y_test_1, verbose=0)
#print(('AWR Test Accuracy: %f' % (CBOW_score2[1]*100)))



# ## F_score Calculation metrix.....................................
# # predict probabilities for test set
# yhat_probs = model.predict(x_test, verbose=0)
# # reduce to 1d array
# yhat_probs = yhat_probs[:, 0]
#
#
# # accuracy: (tp + tn) / (p + n)
# accuracy = accuracy_score(y_test, np.round(abs(yhat_probs)))
# print('CBOW MULti Accuracy: %f' % accuracy)
# # precision tp / (tp + fp)
# precision = precision_score(y_test, np.round(abs(yhat_probs)))
# print('CBOW Multi Precision: %f' % precision)
# # recall: tp / (tp + fn)
# recall = recall_score(y_test, np.round(abs(yhat_probs)))
# print('CBOW Multi Recall: %f' % recall)
# # f1: 2 tp / (2 tp + fp + fn)
# f1 = f1_score(y_test, np.round(abs(yhat_probs)))
# print('CBOW Multi F1 score: %f' % f1)
#
#
# ###CNN CBOW_single model.........................................................
# ##FGFS_labels.......
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# CBOW2_FGFS_labels = []
# for i in range(0, 100):
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in test_FGFS[i]:
#             count += 1
#     if count >= 1:
#         CBOW2_FGFS_labels.append(1)
#     else:
#         CBOW2_FGFS_labels.append(0)
#
#
# #print("FGFS_labels:", FGFS_labels)
#
# print()
# token = Tokenizer()  # create the tokenizer
# token.fit_on_texts(test_FGFS)  # fit the tokenizer on the documents
#
# word_index = token.word_index
# #print('unique words: {}'.format(len(word_index)))
#
#
# max_length = 259
# test = token.texts_to_sequences(test_FGFS)
# x_test = pad_sequences(test, maxlen=max_length, padding='post')
# y_test = np.array(CBOW2_FGFS_labels)
#
#
# ##
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# CBOW2_AWR_labels = []
# for i in range(0, 100):
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in test_AWR[i]:
#             count += 1
#     if count >= 2:
#         CBOW2_AWR_labels.append(1)
#     else:
#         CBOW2_AWR_labels.append(0)
# #print("AWR_labels:", CBOW2_AWR_labels)
#
# print()
# token = Tokenizer()  # create the tokenizer
# token.fit_on_texts(test_AWR)  # fit the tokenizer on the documents
#
# word_index = token.word_index
# #print('unique words: {}'.format(len(word_index)))
#
#
# max_length = 259
# test1 = token.texts_to_sequences(test_AWR)
# x_test_1 = pad_sequences(test1, maxlen=max_length, padding='post')
# y_test_1 = np.array(CBOW2_AWR_labels)
# #print("AWR test label:", y_test_1)
#
# #
# # ##abstract CNN training...................................................
# print("create the tokenizer")
# token = Tokenizer()  # create the tokenizer
# token.fit_on_texts(all_abstract)  # fit the tokenizer on the documents
# #print(token.word_index)
#
#
# word_index = token.word_index
# #print('unique words: {}'.format(len(word_index)))
#
# # # print()
# vocab_size = len(token.word_index) + 1  # define vocabulary size (largest integer value)
# #print('Vocabulary size: %d' % vocab_size)
#
#
# max_length = 259
# train, valid = train_test_split(all_abstract, test_size=0.30, random_state=1)
#
# # print("train", len(train))
# # print("valid", len(valid))
#
# ##train_labels.......................................................................................................
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# CBOW2_train_labels = []
# for i in range(0, 700):
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in all_abstract[i]:
#             count += 1
#     if count >= 1:
#         CBOW2_train_labels.append(1)
#     else:
#         CBOW2_train_labels.append(0)
#
#
# ### validation labels.........................................
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# CBOW2_valid_labels = []
# for i in range(0, 300):
#
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in valid[i]:
#             count += 1
#     if count >= 1:
#         CBOW2_valid_labels.append(1)
#     else:
#         CBOW2_valid_labels.append(0)
#
# #print(CBOW2_valid_labels)
#
#
# train_data = token.texts_to_sequences(train)
# valid_data = token.texts_to_sequences(valid)
# # print("integer incode:", data)
# # print("length: ", len(data))
#
# X_train = pad_sequences(train_data, maxlen=max_length, padding='post')
# y_train = np.array(CBOW2_train_labels)
# #print(len(X_train))
#
# X_valid = pad_sequences(valid_data, max_length, padding='post')
# y_valid = np.array(CBOW2_valid_labels)
#
# bigram_model = Sequential()
# bigram_model.add(Embedding(vocab_size, 100, weights=[embedding_matrix1], input_length=max_length, trainable=False))
# bigram_model.add(Dropout(0.5))
# bigram_model.add(Conv1D(filters=64, kernel_size=2, padding='valid', activation='relu'))
# bigram_model.add(GlobalMaxPool1D())
# bigram_model.add(Flatten())
# bigram_model.add(Dense(10, activation='relu'))
# bigram_model.add(Dropout(0.5))
# bigram_model.add(Dense(1, activation='sigmoid'))
# bigram_model.compile(loss='binary_crossentropy', optimizer='adam', metrics=['accuracy'])
# print(bigram_model.summary())
# plot_model(bigram_model, to_file='bigram_single-channel.png', show_shapes=True, show_layer_names=True)
#
#
# bigram = bigram_model.fit(X_train, y_train,  validation_data=(X_valid, y_valid), epochs=10, verbose=1)
# bigram1 = bigram_model.evaluate(X_train, y_train,  verbose=0)
# print(('Accuracy: %f' % (bigram1[1]*100)))
# bigram2 = bigram_model.evaluate(x_test, y_test, verbose=0)
# print(('FGFS Test Accuracy: %f' % (bigram2[1]*100)))
# bigram3 = bigram_model.evaluate(x_test_1, y_test_1, verbose=0)
# print(('AWR Test Accuracy: %f' % (bigram3[1]*100)))
#
#
# ## F_score Calculation metrix.....................................
# # predict probabilities for test set
# bigram_yhat_probs = bigram_model.predict(x_test, verbose=0)
# # reduce to 1d array
# bigram_yhat_probs = bigram_yhat_probs[:, 0]
#
#
# # accuracy: (tp + tn) / (p + n)
# Bigram_accuracy = accuracy_score(y_test, np.round(abs(bigram_yhat_probs)))
# print('Bigram_Accuracy: %f' % Bigram_accuracy)
# # precision tp / (tp + fp)
# Bigram_precision = precision_score(y_test, np.round(abs(bigram_yhat_probs)))
# print('Bigram_Precision: %f' % Bigram_precision)
# # recall: tp / (tp + fn)
# Bigram_recall = recall_score(y_test, np.round(abs(bigram_yhat_probs)))
# print('Bigram_Recall: %f' % Bigram_recall)
# # f1: 2 tp / (2 tp + fp + fn)
# bigram_f1 = f1_score(y_test, np.round(abs(bigram_yhat_probs)))
# print('bigram_F1_score: %f' % bigram_f1)
#
#
# ###CNN CBOW_single model.........................................................
# ##FGFS_labels.......
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# CBOW3_FGFS_labels = []
# for i in range(0, 100):
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in test_FGFS[i]:
#             count += 1
#     if count >= 2:
#         CBOW3_FGFS_labels.append(1)
#     else:
#         CBOW3_FGFS_labels.append(0)
#
#
# #print("FGFS_labels:", FGFS_labels)
#
# print()
# token = Tokenizer()  # create the tokenizer
# token.fit_on_texts(test_FGFS)  # fit the tokenizer on the documents
#
# word_index = token.word_index
# #print('unique words: {}'.format(len(word_index)))
#
#
# max_length = 259
# test = token.texts_to_sequences(test_FGFS)
# x_test = pad_sequences(test, maxlen=max_length, padding='post')
# y_test = np.array(CBOW3_FGFS_labels)
#
#
# ##
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# CBOW3_AWR_labels = []
# for i in range(0, 100):
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in test_AWR[i]:
#             count += 1
#     if count >= 2:
#         CBOW3_AWR_labels.append(1)
#     else:
#         CBOW3_AWR_labels.append(0)
# #print("AWR_labels:", CBOW2_AWR_labels)
#
# print()
# token = Tokenizer()  # create the tokenizer
# token.fit_on_texts(test_AWR)  # fit the tokenizer on the documents
#
# word_index = token.word_index
# #print('unique words: {}'.format(len(word_index)))
#
#
# max_length = 259
# test1 = token.texts_to_sequences(test_AWR)
# x_test_1 = pad_sequences(test1, maxlen=max_length, padding='post')
# y_test_1 = np.array(CBOW3_AWR_labels)
# #print("AWR test label:", y_test_1)
#
# #
# # ##abstract CNN training...................................................
# print("create the tokenizer")
# token = Tokenizer()  # create the tokenizer
# token.fit_on_texts(all_abstract)  # fit the tokenizer on the documents
# #print(token.word_index)
#
#
# word_index = token.word_index
# #print('unique words: {}'.format(len(word_index)))
#
# # # print()
# vocab_size = len(token.word_index) + 1  # define vocabulary size (largest integer value)
# #print('Vocabulary size: %d' % vocab_size)
#
#
# max_length = 259
# train, valid = train_test_split(all_abstract, test_size=0.30, random_state=1)
#
# # print("train", len(train))
# # print("valid", len(valid))
#
# ##train_labels.......................................................................................................
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# CBOW3_train_labels = []
# for i in range(0, 700):
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in all_abstract[i]:
#             count += 1
#     if count >= 2:
#         CBOW3_train_labels.append(1)
#     else:
#         CBOW3_train_labels.append(0)
#
#
# ### validation labels.........................................
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# CBOW3_valid_labels = []
# for i in range(0, 300):
#
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in valid[i]:
#             count += 1
#     if count >= 2:
#         CBOW3_valid_labels.append(1)
#     else:
#         CBOW3_valid_labels.append(0)
#
# #print(CBOW2_valid_labels)
#
#
# train_data = token.texts_to_sequences(train)
# valid_data = token.texts_to_sequences(valid)
# # print("integer incode:", data)
# # print("length: ", len(data))
#
# X_train = pad_sequences(train_data, maxlen=max_length, padding='post')
# y_train = np.array(CBOW3_train_labels)
# #print(len(X_train))
#
# X_valid = pad_sequences(valid_data, max_length, padding='post')
# y_valid = np.array(CBOW3_valid_labels)
#
# trigram_model = Sequential()
# trigram_model.add(Embedding(vocab_size, 100, weights=[embedding_matrix1], input_length=max_length, trainable=False))
# trigram_model.add(Dropout(0.5))
# trigram_model.add(Conv1D(filters=64, kernel_size=3, padding='valid', activation='relu'))
# trigram_model.add(GlobalMaxPool1D())
# trigram_model.add(Flatten())
# trigram_model.add(Dense(10, activation='relu'))
# trigram_model.add(Dropout(0.5))
# trigram_model.add(Dense(1, activation='sigmoid'))
# trigram_model.compile(loss='binary_crossentropy', optimizer='adam', metrics=['accuracy'])
# print(trigram_model.summary())
#
#
# trigram = trigram_model.fit(X_train, y_train,  validation_data=(X_valid, y_valid), epochs=10, verbose=1)
# trigram1 = trigram_model.evaluate(X_train, y_train,  verbose=0)
# print(('Accuracy: %f' % (trigram1[1]*100)))
# trigram2 = trigram_model.evaluate(x_test, y_test, verbose=0)
# print(('FGFS Test Accuracy: %f' % (trigram2[1]*100)))
# trigram3 = trigram_model.evaluate(x_test_1, y_test_1, verbose=0)
# print(('AWR Test Accuracy: %f' % (trigram3[1]*100)))
#
#
# ## F_score Calculation metrix.....................................
# # predict probabilities for test set
# trigram_yhat_probs = trigram_model.predict(x_test, verbose=0)
# # reduce to 1d array
# trigram_yhat_probs = trigram_yhat_probs[:, 0]
#
#
# # accuracy: (tp + tn) / (p + n)
# Trigram_accuracy = accuracy_score(y_test, np.round(abs(trigram_yhat_probs)))
# print('Trigram_accuracy: %f' % Trigram_accuracy)
# # precision tp / (tp + fp)
# Trigram_precision = precision_score(y_test, np.round(abs(trigram_yhat_probs)))
# print('Trigram_precision: %f' % Trigram_precision)
# # recall: tp / (tp + fn)
# Trigram_recall = recall_score(y_test, np.round(abs(trigram_yhat_probs)))
# print('Trigram_recall: %f' % Trigram_recall)
# # f1: 2 tp / (2 tp + fp + fn)
# Trigram_f1 = f1_score(y_test, np.round(abs(trigram_yhat_probs)))
# print('Trigram_f1_score: %f' % Trigram_f1)
#
#
#
# ###CNN CBOW_single model.........................................................
# ##FGFS_labels.......
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# CBOW4_FGFS_labels = []
# for i in range(0, 100):
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in test_FGFS[i]:
#             count += 1
#     if count >= 3:
#         CBOW4_FGFS_labels.append(1)
#     else:
#         CBOW4_FGFS_labels.append(0)
#
#
# #print("FGFS_labels:", FGFS_labels)
#
# print()
# token = Tokenizer()  # create the tokenizer
# token.fit_on_texts(test_FGFS)  # fit the tokenizer on the documents
#
# word_index = token.word_index
# #print('unique words: {}'.format(len(word_index)))
#
#
# max_length = 259
# test = token.texts_to_sequences(test_FGFS)
# x_test = pad_sequences(test, maxlen=max_length, padding='post')
# y_test = np.array(CBOW4_FGFS_labels)
#
#
# ##
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# CBOW4_AWR_labels = []
# for i in range(0, 100):
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in test_AWR[i]:
#             count += 1
#     if count >= 3:
#         CBOW4_AWR_labels.append(1)
#     else:
#         CBOW4_AWR_labels.append(0)
# #print("AWR_labels:", CBOW2_AWR_labels)
#
# print()
# token = Tokenizer()  # create the tokenizer
# token.fit_on_texts(test_AWR)  # fit the tokenizer on the documents
#
# word_index = token.word_index
# #print('unique words: {}'.format(len(word_index)))
#
#
# max_length = 259
# test1 = token.texts_to_sequences(test_AWR)
# x_test_1 = pad_sequences(test1, maxlen=max_length, padding='post')
# y_test_1 = np.array(CBOW4_AWR_labels)
# #print("AWR test label:", y_test_1)
#
# #
# # ##abstract CNN training...................................................
# print("create the tokenizer")
# token = Tokenizer()  # create the tokenizer
# token.fit_on_texts(all_abstract)  # fit the tokenizer on the documents
# #print(token.word_index)
#
#
# word_index = token.word_index
# #print('unique words: {}'.format(len(word_index)))
#
# # # print()
# vocab_size = len(token.word_index) + 1  # define vocabulary size (largest integer value)
# #print('Vocabulary size: %d' % vocab_size)
#
#
# max_length = 259
# train, valid = train_test_split(all_abstract, test_size=0.30, random_state=1)
#
# # print("train", len(train))
# # print("valid", len(valid))
#
# ##train_labels.......................................................................................................
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# CBOW4_train_labels = []
# for i in range(0, 700):
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in all_abstract[i]:
#             count += 1
#     if count >= 3:
#         CBOW4_train_labels.append(1)
#     else:
#         CBOW4_train_labels.append(0)
#
#
# ### validation labels.........................................
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# CBOW4_valid_labels = []
# for i in range(0, 300):
#
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in valid[i]:
#             count += 1
#     if count >= 3:
#         CBOW4_valid_labels.append(1)
#     else:
#         CBOW4_valid_labels.append(0)
#
# #print(CBOW2_valid_labels)
#
#
# train_data = token.texts_to_sequences(train)
# valid_data = token.texts_to_sequences(valid)
# # print("integer incode:", data)
# # print("length: ", len(data))
#
# X_train = pad_sequences(train_data, maxlen=max_length, padding='post')
# y_train = np.array(CBOW4_train_labels)
# #print(len(X_train))
#
# X_valid = pad_sequences(valid_data, max_length, padding='post')
# y_valid = np.array(CBOW4_valid_labels)
#
# fourgram_model = Sequential()
# fourgram_model.add(Embedding(vocab_size, 100, weights=[embedding_matrix1], input_length=max_length, trainable=False))
# fourgram_model.add(Dropout(0.5))
# fourgram_model.add(Conv1D(filters=64, kernel_size=3, padding='valid', activation='relu'))
# fourgram_model.add(GlobalMaxPool1D())
# fourgram_model.add(Flatten())
# fourgram_model.add(Dense(10, activation='relu'))
# fourgram_model.add(Dropout(0.5))
# fourgram_model.add(Dense(1, activation='sigmoid'))
# fourgram_model.compile(loss='binary_crossentropy', optimizer='adam', metrics=['accuracy'])
# print(fourgram_model.summary())
#
#
# fourgram = fourgram_model.fit(X_train, y_train,  validation_data=(X_valid, y_valid), epochs=10, verbose=1)
# fourgram1 = fourgram_model.evaluate(X_train, y_train,  verbose=0)
# print(('Accuracy: %f' % (fourgram1[1]*100)))
# fourgram2 = fourgram_model.evaluate(x_test, y_test, verbose=0)
# print(('FGFS Test Accuracy: %f' % (fourgram2[1]*100)))
# fourgram3 = fourgram_model.evaluate(x_test_1, y_test_1, verbose=0)
# print(('AWR Test Accuracy: %f' % (fourgram3[1]*100)))
#
#
# ## F_score Calculation metrix.....................................
# # predict probabilities for test set
# Fourgram_yhat_probs = fourgram_model.predict(x_test, verbose=0)
# # reduce to 1d array
# Fourgram_yhat_probs = Fourgram_yhat_probs[:, 0]
#
#
# # accuracy: (tp + tn) / (p + n)
# Fourgram_accuracy = accuracy_score(y_test, np.round(abs(Fourgram_yhat_probs)))
# print('Fourgram_accuracy: %f' % Fourgram_accuracy)
# # precision tp / (tp + fp)
# Fourgram_precision = precision_score(y_test, np.round(abs(Fourgram_yhat_probs)))
# print('Fourgram_precision: %f' % Fourgram_precision)
# # recall: tp / (tp + fn)
# Fourgram_recall = recall_score(y_test, np.round(abs(Fourgram_yhat_probs)))
# print('Fourgram_recall: %f' % Fourgram_recall)
# # f1: 2 tp / (2 tp + fp + fn)
# Fourgram_f1 = f1_score(y_test, np.round(abs(Fourgram_yhat_probs)))
# print('Fourgram_f1_score: %f' % Fourgram_f1)
#
#
# ###plot accuracy during training..............................................................
# pyplot.plot(bigram.history['accuracy'], label='CBOW_bigram_train')
# #pyplot.plot(bigram.history['val_accuracy'], label='CBOW_bigram_valid')
# pyplot.plot(trigram.history['accuracy'], label='CBOW_trigram_train')
# #pyplot.plot(trigram.history['val_accuracy'], label='CBOW_trigram_valid')
# pyplot.plot(fourgram.history['accuracy'], label='CBOW_fourgram_train')
# #pyplot.plot(fourgram.history['val_accuracy'], label='CBOW_fourgram_valid')
# # pyplot.plot(CBOW_multi.history['loss'], label='train_loss')
# # pyplot.plot(CBOW_multi.history['val_loss'], label='val_loss')
# pyplot.title('CBOW Embedding Single-Channel CNN Models')
# pyplot.xlabel('Epoch')
# pyplot.ylabel('Accuracy')
# pyplot.grid(True)
# pyplot.legend(['CBOW_bigram_train', 'CBOW_trigram_train', 'CBOW_fourgram_train',], loc='lower right')
# pyplot.show()



#
#
# ##pre-traind word2vec data.......................................................
# sg_embeddings = Word2Vec(sentences=all_abstract, size=100, min_count=1)
# sg_embeddings.wv.save_word2vec_format('Sg_Pre-trained_word2Vec.txt', binary=False)
#
# ##FGFS_labels..................................................................
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# Sg_FGFS_labels = []
# for i in range(0, 100):
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in test_FGFS[i]:
#             count += 1
#     if count >= 2:
#         Sg_FGFS_labels.append(1)
#     else:
#         Sg_FGFS_labels.append(0)
#
#
# #print("FGFS_labels:", FGFS_labels)
#
# print()
# token = Tokenizer()  # create the tokenizer
# token.fit_on_texts(test_FGFS)  # fit the tokenizer on the documents
#
# word_index = token.word_index
# #print('unique words: {}'.format(len(word_index)))
#
#
# max_length = 259
# test = token.texts_to_sequences(test_FGFS)
# x_test = pad_sequences(test, maxlen=max_length, padding='post')
# y_test = np.array(Sg_FGFS_labels)
#
#
#
#
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# Sg_AWR_labels = []
# for i in range(0, 100):
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in test_AWR[i]:
#             count += 1
#     if count >= 2:
#         Sg_AWR_labels.append(1)
#     else:
#         Sg_AWR_labels.append(0)
# print("AWR_labels:", Sg_AWR_labels)
#
# print()
# token = Tokenizer()  # create the tokenizer
# token.fit_on_texts(test_AWR)  # fit the tokenizer on the documents
#
# word_index = token.word_index
# #print('unique words: {}'.format(len(word_index)))
#
#
# max_length = 259
# test1 = token.texts_to_sequences(test_AWR)
# x_test_1 = pad_sequences(test1, maxlen=max_length, padding='post')
# y_test_1 = np.array(Sg_AWR_labels)
# print("AWR test label:", y_test_1)
#
#
# # ##abstract CNN training...................................................
# print("create the tokenizer")
# token = Tokenizer()  # create the tokenizer
# token.fit_on_texts(all_abstract)  # fit the tokenizer on the documents
# #print(token.word_index)
#
#
# word_index = token.word_index
# #print('unique words: {}'.format(len(word_index)))
#
# # # print()
# vocab_size = len(token.word_index) + 1  # define vocabulary size (largest integer value)
# #print('Vocabulary size: %d' % vocab_size)
#
#
# max_length = 259
# train, valid = train_test_split(all_abstract, test_size=0.30, random_state=1)
# #print("train", len(train))
# #print("valid", len(valid))
#
# ##train_labels.......................................................................................................
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# Sg_train_labels = []
# for i in range(0, 700):
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in all_abstract[i]:
#             count += 1
#     if count >= 6:
#         Sg_train_labels.append(1)
#     else:
#         Sg_train_labels.append(0)
#
#
# ### validation labels.........................................
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# Sg_valid_labels = []
# for i in range(0, 300):
#
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in valid[i]:
#             count += 1
#     if count >= 1:
#         Sg_valid_labels.append(1)
#     else:
#         Sg_valid_labels.append(0)
# #print(valid_labels)
#
#
# train_data = token.texts_to_sequences(train)
# valid_data = token.texts_to_sequences(valid)
#
#
# X_train = pad_sequences(train_data, maxlen=max_length, padding='post')
# y_train = np.array(Sg_train_labels)
# #print(len(X_train))
#
# X_valid = pad_sequences(valid_data, max_length, padding='post')
# y_valid = np.array(Sg_valid_labels)
#
#
# ###Sg_CNNs model with word2vec ....................................................................
# sg_embedding_index = {}
# list_v = []
# file = open('Sg_Pre-trained_word2Vec.txt', 'rt', encoding='UTF8')
# line = file.readline()
# totalWords, numOfFeatures = line.split()
# print(totalWords, numOfFeatures)
# for line in file:
#     values = line.split()
#     list_v.append(values)
#     word = values[0]
#     coefs = array(values[1:], dtype='float64')
#     sg_embedding_index[word] = coefs
#
#
# print('Found %s word vectors.' % len(sg_embedding_index))
# df_values = pd.DataFrame(list_v)
# print(df_values, "\n")
#
# sg_embedding_matrix1 = np.array([[0 for col in range(100)] for row in range(4787)])
# for word, i in token.word_index.items():
#     # try:
#     embedding_vector = sg_embedding_index.get(word)
#     if embedding_vector is not None:
#         if( i == 100):
#             print(i,"번째 완료")
#         for j in range(0, 100):
#            sg_embedding_matrix1[i][j] = embedding_vector[j]
#         #print(i,"번째 완료")
#
# #print("sg_embedding_matrix1:", sg_embedding_matrix1)
#
#
# ##Sg_CNNs word2vec model ................................................................................
# embedding_dim = 100
#
# Sg_model = Sequential()
# inputs1 = Input(shape=(max_length,))
# embedding1 = Embedding(vocab_size, 100, weights=[sg_embedding_matrix1], input_length=max_length, trainable=False)(inputs1)
# conv1 = Conv1D(filters=32, kernel_size=2, activation='relu')(embedding1)
# drop1 = Dropout(0.3)(conv1)
# pool1 = GlobalMaxPool1D()(drop1)
# flat1 = Flatten()(pool1)
# #channel 2
# conv2 = Conv1D(filters=32, kernel_size=3, activation='relu')(embedding1)
# drop2 = Dropout(0.3)(conv2)
# pool2 = GlobalMaxPool1D()(drop2)
# flat2 = Flatten()(pool2)
# #channel 3
# conv3 = Conv1D(filters=32, kernel_size=4, activation='relu')(embedding1)
# drop3 = Dropout(0.3)(conv3)
# pool3 = GlobalMaxPool1D()(drop3)
# flat3 = Flatten()(pool3)
# #merge
# merged = concatenate([flat1, flat2, flat3])
# #interpretation
# dense1 = Dense(10, activation='relu')(merged)
# outputs = Dense(1, activation='sigmoid')(dense1)
#
# Sg_model = Model(inputs=[inputs1], outputs=outputs)
# # compile
# Sg_model.compile(loss='binary_crossentropy', optimizer='adam', metrics=['acc'])
# # summarize
# print(Sg_model.summary())
# plot_model(Sg_model, to_file='Word2Vec_Multichannel.png', show_shapes=True, show_layer_names=True)
# pyplot.show()
#
#
# Sg_history = Sg_model.fit(X_train, y_train,  validation_data=(X_valid, y_valid), epochs=100, verbose=1)
# Sg_score = Sg_model.evaluate(X_train, y_train,  verbose=0)
# print(('Accuracy: %f' % (Sg_score[1]*100)))
# Sg_score1 = Sg_model.evaluate(x_test, y_test, verbose=0)
# print(('FGFS Test Accuracy: %f' % (Sg_score1[1]*100)))
# Sg_score2 = Sg_model.evaluate(x_test_1, y_test_1, verbose=0)
# print(('AWR Test Accuracy: %f' % (Sg_score2[1]*100)))
#
#
# ###plot accuracy during training..............................................................
# pyplot.plot(CBOW_multi.history['acc'], label='acc')
# pyplot.plot(CBOW_multi.history['val_acc'], label='val_acc')
# pyplot.plot(Sg_history.history['acc'], label='acc')
# pyplot.plot(Sg_history.history['val_acc'], label='val_acc')
# pyplot.title('CBOW & Sg Embedding Multi-Channel CNN Models')
# pyplot.xlabel('Epoch')
# pyplot.ylabel('Accuracy')
# pyplot.grid(True)
# pyplot.legend(['CBOW_Acc', 'CBOW_Val_Acc', 'Sg_Acc', 'Sg_Val_Acc'], loc='lower right')
# pyplot.show()
#

#
# ## F_score Calculation metrix.....................................
# # predict probabilities for test set
# Sg_multi_yhat_probs = Sg_model.predict(x_test, verbose=0)
# # reduce to 1d array
# Sg_multi_yhat_probs = Sg_multi_yhat_probs[:, 0]
#
#
# # accuracy: (tp + tn) / (p + n)
# Sg_multi_accuracy = accuracy_score(y_test, np.round(abs(Sg_multi_yhat_probs)))
# print('Sg_multi_accuracy: %f' % Sg_multi_accuracy)
# # precision tp / (tp + fp)
# Sg_multi_precision = precision_score(y_test, np.round(abs(Sg_multi_yhat_probs)))
# print('Sg_multi_precision: %f' % Sg_multi_precision)
# # recall: tp / (tp + fn)
# Sg_multi_recall = recall_score(y_test, np.round(abs(Sg_multi_yhat_probs)))
# print('Sg_multi_recall: %f' % Sg_multi_recall)
# # f1: 2 tp / (2 tp + fp + fn)
# Sg_f1 = f1_score(y_test, np.round(abs(Sg_multi_yhat_probs)))
# print('Sg_multi_f1_score: %f' % Sg_f1)
#
#
#
# ####CNN CBOW_single model.........................................................
# ##FGFS_labels.......
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# Sg2_FGFS_labels = []
# for i in range(0, 100):
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in test_FGFS[i]:
#             count += 1
#     if count >= 3:
#         Sg2_FGFS_labels.append(1)
#     else:
#         Sg2_FGFS_labels.append(0)
#
#
# #print("FGFS_labels:", FGFS_labels)
#
# print()
# token = Tokenizer()  # create the tokenizer
# token.fit_on_texts(test_FGFS)  # fit the tokenizer on the documents
#
# word_index = token.word_index
# #print('unique words: {}'.format(len(word_index)))
#
#
# max_length = 259
# test = token.texts_to_sequences(test_FGFS)
# x_test = pad_sequences(test, maxlen=max_length, padding='post')
# y_test = np.array(Sg2_FGFS_labels)
#
#
# ##
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# Sg2_AWR_labels = []
# for i in range(0, 100):
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in test_AWR[i]:
#             count += 1
#     if count >= 1:
#         Sg2_AWR_labels.append(1)
#     else:
#         Sg2_AWR_labels.append(0)
# #print("AWR_labels:", CBOW2_AWR_labels)
#
# print()
# token = Tokenizer()  # create the tokenizer
# token.fit_on_texts(test_AWR)  # fit the tokenizer on the documents
#
# word_index = token.word_index
# #print('unique words: {}'.format(len(word_index)))
#
#
# max_length = 259
# test1 = token.texts_to_sequences(test_AWR)
# x_test_1 = pad_sequences(test1, maxlen=max_length, padding='post')
# y_test_1 = np.array(Sg2_AWR_labels)
# #print("AWR test label:", y_test_1)
#
# #
# # ##abstract CNN training...................................................
# print("create the tokenizer")
# token = Tokenizer()  # create the tokenizer
# token.fit_on_texts(all_abstract)  # fit the tokenizer on the documents
# #print(token.word_index)
#
#
# word_index = token.word_index
# #print('unique words: {}'.format(len(word_index)))
#
# # # print()
# vocab_size = len(token.word_index) + 1  # define vocabulary size (largest integer value)
# #print('Vocabulary size: %d' % vocab_size)
#
#
# max_length = 259
# train, valid = train_test_split(all_abstract, test_size=0.30, random_state=1)
#
# # print("train", len(train))
# # print("valid", len(valid))
#
# ##train_labels.......................................................................................................
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# Sg2_train_labels = []
# for i in range(0, 700):
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in all_abstract[i]:
#             count += 1
#     if count >= 3:
#         Sg2_train_labels.append(1)
#     else:
#         Sg2_train_labels.append(0)
#
#
# ### validation labels.........................................
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# Sg2_valid_labels = []
# for i in range(0, 300):
#
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in valid[i]:
#             count += 1
#     if count >= 3:
#         Sg2_valid_labels.append(1)
#     else:
#         Sg2_valid_labels.append(0)
#
# #print(CBOW2_valid_labels)
#
#
# train_data = token.texts_to_sequences(train)
# valid_data = token.texts_to_sequences(valid)
# # print("integer incode:", data)
# # print("length: ", len(data))
#
# X_train = pad_sequences(train_data, maxlen=max_length, padding='post')
# y_train = np.array(Sg2_train_labels)
# #print(len(X_train))
#
# X_valid = pad_sequences(valid_data, max_length, padding='post')
# y_valid = np.array(Sg2_valid_labels)
#
# Sg_bigram_model = Sequential()
# Sg_bigram_model.add(Embedding(vocab_size, 100, weights=[sg_embedding_matrix1], input_length=max_length, trainable=False))
# Sg_bigram_model.add(Dropout(0.5))
# Sg_bigram_model.add(Conv1D(filters=64, kernel_size=2, padding='valid', activation='relu'))
# Sg_bigram_model.add(GlobalMaxPool1D())
# Sg_bigram_model.add(Flatten())
# Sg_bigram_model.add(Dense(10, activation='relu'))
# Sg_bigram_model.add(Dropout(0.5))
# Sg_bigram_model.add(Dense(1, activation='sigmoid'))
# Sg_bigram_model.compile(loss='binary_crossentropy', optimizer='adam', metrics=['accuracy'])
# print(Sg_bigram_model.summary())
#
#
# Sg_bigram = Sg_bigram_model.fit(X_train, y_train,  validation_data=(X_valid, y_valid), epochs=10, verbose=1)
# Sg_bigram1 = Sg_bigram_model.evaluate(X_train, y_train,  verbose=0)
# print(('Accuracy: %f' % (Sg_bigram1[1]*100)))
# Sg_bigram2 = Sg_bigram_model.evaluate(x_test, y_test, verbose=0)
# print(('FGFS Test Accuracy: %f' % (Sg_bigram2[1]*100)))
# Sg_bigram3 = Sg_bigram_model.evaluate(x_test_1, y_test_1, verbose=0)
# print(('AWR Test Accuracy: %f' % (Sg_bigram3[1]*100)))
#
#
#
# ## F_score Calculation metrix.....................................
# # predict probabilities for test set
# Sg_bigram_yhat_probs = Sg_bigram_model.predict(x_test, verbose=0)
# # reduce to 1d array
# Sg_bigram_yhat_probs = Sg_bigram_yhat_probs[:, 0]
#
#
# # accuracy: (tp + tn) / (p + n)
# Sg_bigram_accuracy = accuracy_score(y_test, np.round(abs(Sg_bigram_yhat_probs)))
# print('Sg_bigram_accuracy: %f' % Sg_bigram_accuracy)
# # precision tp / (tp + fp)
# Sg_bigram_precision = precision_score(y_test, np.round(abs(Sg_bigram_yhat_probs)))
# print('Sg_bigram_precision: %f' % Sg_bigram_precision)
# # recall: tp / (tp + fn)
# Sg_bigram_recall = recall_score(y_test, np.round(abs(Sg_bigram_yhat_probs)))
# print('Sg_bigram_recall: %f' % Sg_bigram_recall)
# # f1: 2 tp / (2 tp + fp + fn)
# Sg_bigram_f1 = f1_score(y_test, np.round(abs(Sg_bigram_yhat_probs)))
# print('Sg_bigram_f1_score: %f' % Sg_bigram_f1)
#
#
#
# ###CNN CBOW_single model.........................................................
# ##FGFS_labels.......
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# Sg3_FGFS_labels = []
# for i in range(0, 100):
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in test_FGFS[i]:
#             count += 1
#     if count >= 2:
#         Sg3_FGFS_labels.append(1)
#     else:
#         Sg3_FGFS_labels.append(0)
#
#
# #print("FGFS_labels:", FGFS_labels)
#
# print()
# token = Tokenizer()  # create the tokenizer
# token.fit_on_texts(test_FGFS)  # fit the tokenizer on the documents
#
# word_index = token.word_index
# #print('unique words: {}'.format(len(word_index)))
#
#
# max_length = 259
# test = token.texts_to_sequences(test_FGFS)
# x_test = pad_sequences(test, maxlen=max_length, padding='post')
# y_test = np.array(Sg3_FGFS_labels)
#
#
# ##
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# Sg3_AWR_labels = []
# for i in range(0, 100):
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in test_AWR[i]:
#             count += 1
#     if count >= 2:
#         Sg3_AWR_labels.append(1)
#     else:
#         Sg3_AWR_labels.append(0)
# #print("AWR_labels:", CBOW2_AWR_labels)
#
# print()
# token = Tokenizer()  # create the tokenizer
# token.fit_on_texts(test_AWR)  # fit the tokenizer on the documents
#
# word_index = token.word_index
# #print('unique words: {}'.format(len(word_index)))
#
#
# max_length = 259
# test1 = token.texts_to_sequences(test_AWR)
# x_test_1 = pad_sequences(test1, maxlen=max_length, padding='post')
# y_test_1 = np.array(Sg3_AWR_labels)
# #print("AWR test label:", y_test_1)
#
# #
# # ##abstract CNN training...................................................
# print("create the tokenizer")
# token = Tokenizer()  # create the tokenizer
# token.fit_on_texts(all_abstract)  # fit the tokenizer on the documents
# #print(token.word_index)
#
#
# word_index = token.word_index
# #print('unique words: {}'.format(len(word_index)))
#
# # # print()
# vocab_size = len(token.word_index) + 1  # define vocabulary size (largest integer value)
# #print('Vocabulary size: %d' % vocab_size)
#
#
# max_length = 259
# train, valid = train_test_split(all_abstract, test_size=0.30, random_state=1)
#
# # print("train", len(train))
# # print("valid", len(valid))
#
# ##train_labels.......................................................................................................
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# Sg3_train_labels = []
# for i in range(0, 700):
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in all_abstract[i]:
#             count += 1
#     if count >= 2:
#         Sg3_train_labels.append(1)
#     else:
#         Sg3_train_labels.append(0)
#
#
# ### validation labels.........................................
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# Sg3_valid_labels = []
# for i in range(0, 300):
#
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in valid[i]:
#             count += 1
#     if count >= 2:
#         Sg3_valid_labels.append(1)
#     else:
#         Sg3_valid_labels.append(0)
#
# #print(CBOW2_valid_labels)
#
#
# train_data = token.texts_to_sequences(train)
# valid_data = token.texts_to_sequences(valid)
# # print("integer incode:", data)
# # print("length: ", len(data))
#
# X_train = pad_sequences(train_data, maxlen=max_length, padding='post')
# y_train = np.array(Sg3_train_labels)
# #print(len(X_train))
#
# X_valid = pad_sequences(valid_data, max_length, padding='post')
# y_valid = np.array(Sg3_valid_labels)
#
# Sg_trigram_model = Sequential()
# Sg_trigram_model.add(Embedding(vocab_size, 100, weights=[sg_embedding_matrix1], input_length=max_length, trainable=False))
# Sg_trigram_model.add(Dropout(0.5))
# Sg_trigram_model.add(Conv1D(filters=64, kernel_size=3, padding='valid', activation='relu'))
# Sg_trigram_model.add(GlobalMaxPool1D())
# Sg_trigram_model.add(Flatten())
# Sg_trigram_model.add(Dense(10, activation='relu'))
# Sg_trigram_model.add(Dropout(0.5))
# Sg_trigram_model.add(Dense(1, activation='sigmoid'))
# Sg_trigram_model.compile(loss='binary_crossentropy', optimizer='adam', metrics=['accuracy'])
# print(Sg_trigram_model.summary())
#
#
# Sg_trigram = Sg_trigram_model.fit(X_train, y_train,  validation_data=(X_valid, y_valid), epochs=10, verbose=1)
# Sg_trigram1 = Sg_trigram_model.evaluate(X_train, y_train,  verbose=0)
# print(('Accuracy: %f' % (Sg_trigram1[1]*100)))
# Sg_trigram2 = Sg_trigram_model.evaluate(x_test, y_test, verbose=0)
# print(('FGFS Test Accuracy: %f' % (Sg_trigram2[1]*100)))
# Sg_trigram3 = Sg_trigram_model.evaluate(x_test_1, y_test_1, verbose=0)
# print(('AWR Test Accuracy: %f' % (Sg_trigram3[1]*100)))
#
#
#
# ## F_score Calculation metrix.....................................
# # predict probabilities for test set
# Sg_trigram_yhat_probs = Sg_trigram_model.predict(x_test, verbose=0)
# # reduce to 1d array
# Sg_trigram_yhat_probs = Sg_trigram_yhat_probs[:, 0]
#
#
# # accuracy: (tp + tn) / (p + n)
# Sg_trigram_accuracy = accuracy_score(y_test, np.round(abs(Sg_trigram_yhat_probs)))
# print('Sg_trigram_accuracy: %f' % Sg_trigram_accuracy)
# # precision tp / (tp + fp)
# Sg_trigram_precision = precision_score(y_test, np.round(abs(Sg_trigram_yhat_probs)))
# print('Sg_trigram_precision: %f' % Sg_trigram_precision)
# # recall: tp / (tp + fn)
# Sg_trigram_recall = recall_score(y_test, np.round(abs(Sg_trigram_yhat_probs)))
# print('Sg_trigram_recall: %f' % Sg_trigram_recall)
# # f1: 2 tp / (2 tp + fp + fn)
# Sg_trigram_f1 = f1_score(y_test, np.round(abs(Sg_trigram_yhat_probs)))
# print('Sg_trigram_f1_score: %f' % Sg_trigram_f1)
#
#
# ###CNN CBOW_single model.........................................................
# ##FGFS_labels.......
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# Sg4_FGFS_labels = []
# for i in range(0, 100):
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in test_FGFS[i]:
#             count += 1
#     if count >= 1:
#         Sg4_FGFS_labels.append(1)
#     else:
#         Sg4_FGFS_labels.append(0)
#
#
# #print("FGFS_labels:", FGFS_labels)
#
# print()
# token = Tokenizer()  # create the tokenizer
# token.fit_on_texts(test_FGFS)  # fit the tokenizer on the documents
#
# word_index = token.word_index
# #print('unique words: {}'.format(len(word_index)))
#
#
# max_length = 259
# test = token.texts_to_sequences(test_FGFS)
# x_test = pad_sequences(test, maxlen=max_length, padding='post')
# y_test = np.array(Sg4_FGFS_labels)
#
#
# ##
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# Sg4_AWR_labels = []
# for i in range(0, 100):
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in test_AWR[i]:
#             count += 1
#     if count >= 1:
#         Sg4_AWR_labels.append(1)
#     else:
#         Sg4_AWR_labels.append(0)
# #print("AWR_labels:", CBOW2_AWR_labels)
#
# print()
# token = Tokenizer()  # create the tokenizer
# token.fit_on_texts(test_AWR)  # fit the tokenizer on the documents
#
# word_index = token.word_index
# #print('unique words: {}'.format(len(word_index)))
#
#
# max_length = 259
# test1 = token.texts_to_sequences(test_AWR)
# x_test_1 = pad_sequences(test1, maxlen=max_length, padding='post')
# y_test_1 = np.array(Sg4_AWR_labels)
# #print("AWR test label:", y_test_1)
#
# #
# # ##abstract CNN training...................................................
# print("create the tokenizer")
# token = Tokenizer()  # create the tokenizer
# token.fit_on_texts(all_abstract)  # fit the tokenizer on the documents
# #print(token.word_index)
#
#
# word_index = token.word_index
# #print('unique words: {}'.format(len(word_index)))
#
# # # print()
# vocab_size = len(token.word_index) + 1  # define vocabulary size (largest integer value)
# #print('Vocabulary size: %d' % vocab_size)
#
#
# max_length = 259
# train, valid = train_test_split(all_abstract, test_size=0.30, random_state=1)
#
# # print("train", len(train))
# # print("valid", len(valid))
#
# ##train_labels.......................................................................................................
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# Sg4_train_labels = []
# for i in range(0, 700):
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in all_abstract[i]:
#             count += 1
#     if count >= 1:
#         Sg4_train_labels.append(1)
#     else:
#         Sg4_train_labels.append(0)
#
#
# ### validation labels.........................................
# select_words = ['system',  'network', 'approach', 'time', 'cloud', 'information', 'process', 'problem', 'service', 'security']
# Sg4_valid_labels = []
# for i in range(0, 300):
#
#     count = 0
#     for j in range(0, len(select_words)):
#         if select_words[j] in valid[i]:
#             count += 1
#     if count >= 1:
#         Sg4_valid_labels.append(1)
#     else:
#         Sg4_valid_labels.append(0)
#
# #print(CBOW2_valid_labels)
#
#
# train_data = token.texts_to_sequences(train)
# valid_data = token.texts_to_sequences(valid)
# # print("integer incode:", data)
# # print("length: ", len(data))
#
# X_train = pad_sequences(train_data, maxlen=max_length, padding='post')
# y_train = np.array(Sg4_train_labels)
# #print(len(X_train))
#
# X_valid = pad_sequences(valid_data, max_length, padding='post')
# y_valid = np.array(Sg4_valid_labels)
#
# Sg_fourgram_model = Sequential()
# Sg_fourgram_model.add(Embedding(vocab_size, 100, weights=[sg_embedding_matrix1], input_length=max_length, trainable=False))
# Sg_fourgram_model.add(Dropout(0.5))
# Sg_fourgram_model.add(Conv1D(filters=64, kernel_size=3, padding='valid', activation='relu'))
# Sg_fourgram_model.add(GlobalMaxPool1D())
# Sg_fourgram_model.add(Flatten())
# Sg_fourgram_model.add(Dense(10, activation='relu'))
# Sg_fourgram_model.add(Dropout(0.5))
# Sg_fourgram_model.add(Dense(1, activation='sigmoid'))
# Sg_fourgram_model.compile(loss='binary_crossentropy', optimizer='adam', metrics=['accuracy'])
# print(Sg_fourgram_model.summary())
#
#
# Sg_fourgram = Sg_fourgram_model.fit(X_train, y_train,  validation_data=(X_valid, y_valid), epochs=10, verbose=1)
# Sg_fourgram1 = Sg_fourgram_model.evaluate(X_train, y_train,  verbose=0)
# print(('Accuracy: %f' % (Sg_fourgram1[1]*100)))
# Sg_fourgram2 = Sg_fourgram_model.evaluate(x_test, y_test, verbose=0)
# print(('FGFS Test Accuracy: %f' % (Sg_fourgram2[1]*100)))
# Sg_fourgram3 = Sg_fourgram_model.evaluate(x_test_1, y_test_1, verbose=0)
# print(('AWR Test Accuracy: %f' % (Sg_fourgram3[1]*100)))
#
#
#
# ## F_score Calculation metrix.....................................
# # predict probabilities for test set
# Sg_fourgram_yhat_probs = Sg_fourgram_model.predict(x_test, verbose=0)
# # reduce to 1d array
# Sg_fourgram_yhat_probs = Sg_fourgram_yhat_probs[:, 0]
#
#
# # accuracy: (tp + tn) / (p + n)
# Sg_fourgram_accuracy = accuracy_score(y_test, np.round(abs(Sg_fourgram_yhat_probs)))
# print('Sg_fourgram_accuracy: %f' % Sg_fourgram_accuracy)
# # precision tp / (tp + fp)
# Sg_fourgram_precision = precision_score(y_test, np.round(abs(Sg_fourgram_yhat_probs)))
# print('Sg_fourgram_precision: %f' % Sg_fourgram_precision)
# # recall: tp / (tp + fn)
# Sg_fourgram_recall = recall_score(y_test, np.round(abs(Sg_fourgram_yhat_probs)))
# print('Sg_fourgram_recall: %f' % Sg_fourgram_recall)
# # f1: 2 tp / (2 tp + fp + fn)
# Sg_fourgram_f1 = f1_score(y_test, np.round(abs(Sg_fourgram_yhat_probs)))
# print('Sg_fourgram_f1_score: %f' % Sg_fourgram_f1)
#
#
# ###plot accuracy during training.................................................
# pyplot.plot(Sg_bigram.history['accuracy'], label='Sg_bigram_train')
# #pyplot.plot(Sg_bigram.history['val_accuracy'], label='Sg_bigram_valid')
# pyplot.plot(Sg_trigram.history['accuracy'], label='Sg_trigram_train')
# #pyplot.plot(Sg_trigram.history['val_accuracy'], label='Sg_trigram_valid')
# pyplot.plot(Sg_fourgram.history['accuracy'], label='Sg_fourgram_train')
# #pyplot.plot(Sg_fourgram.history['val_accuracy'], label='Sg_fourgram_valid')
# # pyplot.plot(CBOW_multi.history['loss'], label='train_loss')
# # pyplot.plot(CBOW_multi.history['val_loss'], label='val_loss')
# pyplot.title('Sg Embedding Single-Channel CNN Models')
# pyplot.xlabel('Epoch')
# pyplot.ylabel('Accuracy')
# pyplot.grid(True)
# pyplot.legend(['Sg_bigram_train', 'Sg_trigram_train', 'Sg_fourgram_train'], loc='lower right')
# pyplot.show()
#


#
# ##Sg Single Traing & Test 3d graph drawing..............................
# x = np.array(range(3), float)
# y = x.copy()
# xpos, ypos = np.meshgrid(x, y)
# z = np.array([[96.7999, 94.1428, 72.0003],
#                [75.0000, 72.0003, 47.9999],
#                [62.9999, 52.9999, 26.0000]])
#
# color = ['b', 'g', 'crimson']*3
# xpos = xpos.flatten()
# ypos = ypos.flatten()
# zpos = np.zeros_like(xpos)
# dx = 0.5*np.ones_like(zpos)
# dy = dx.copy()
# dz = z.flatten()
#
# fig = plt.figure()
# ax = fig.add_subplot(111, projection = '3d')
# labels = ["Sg_Bigram", "Sg_Trigram", "Sg_Fourgram"]
# ax.set_title("Sg Single Models Accuracy", fontsize=13)
# ax.set_xticklabels(labels, fontstyle='italic')
# ax.set_xticks(range(3))
# #ax.set_xlabel('Sg Models', labelpad=7, fontweight='bold')
#
# ax.set_yticks(range(3))
# ax.set_yticklabels(["FGFS_Train", "FGFS_Test", "AWR_Test"], fontstyle='italic')
# ax.set_ylabel("Train & Test", labelpad=10, fontstyle='italic', fontweight='bold')
#
# ax.set_zticklabels([0, 20, 40, 60, 80, 100])
# ax.set_zlabel("Accuracy")
#
# x_leg = plt.Rectangle((0, 0), 1, 1, fc='b')
# y_leg = plt.Rectangle((0, 0), 1, 1, fc='g')
# z_leg = plt.Rectangle((0, 0), 1, 1, fc='crimson')
# ax.legend([x_leg, y_leg, z_leg], labels, fontsize=10, loc='upper right')
#
# ax.bar3d(xpos, ypos, zpos, dx, dy, dz, color=color)
# plt.show()
#
#
#
# ##CBOW Single Train & Test 3d graph drawing..............................
# x = np.array(range(3), float)
# y = x.copy()
# xpos, ypos = np.meshgrid(x, y)
# z = np.array([[94.1428, 93.0001, 69.9999],
#               [80.2857, 75.0000, 62.0000],
#               [72.0000, 62.9999, 34.0000]])
#
# color = ['c', 'm', 'y']*3
# xpos = xpos.flatten()
# ypos = ypos.flatten()
# zpos = np.zeros_like(xpos)
# dx = 0.5*np.ones_like(zpos)
# dy = dx.copy()
# dz = z.flatten()
#
# fig = plt.figure()
# ax = fig.add_subplot(111, projection = '3d')
# labels = ["CBOW_Bigram", "CBOW_Trigram", "CBOW_Fourgram"]
# ax.set_title("CBOW Single Models Accuracy", fontsize=13)
# ax.set_xticklabels(labels, fontstyle='italic')
# ax.set_xticks(range(3))
# #ax.set_xlabel('CBOW Models', labelpad=7, fontweight='bold')
#
# ax.set_yticks(range(3))
# ax.set_yticklabels(["FGFS_Train", "FGFS_Test", "AWR_Test"], fontstyle='italic')
# ax.set_ylabel("Train & Test", labelpad=10, fontstyle='italic', fontweight='bold')
#
# ax.set_zticklabels([0, 20, 40, 60, 80, 100])
# ax.set_zlabel("Accuracy")
#
# x_leg = plt.Rectangle((0, 0), 1, 1, fc="c")
# y_leg = plt.Rectangle((0, 0), 1, 1, fc="m")
# z_leg = plt.Rectangle((0, 0), 1, 1, fc="y")
# ax.legend([x_leg, y_leg, z_leg], labels, fontsize=10, loc='upper right')
#
# ax.bar3d(xpos, ypos, zpos, dx, dy, dz, color=color)
# plt.show()




# # Test_evaluation......................................
# # x = np.linspace(0, 5, 100)
# # pyplot.plot(x, x**2+0.8927, c='b', ls='-')
# # pyplot.plot(x, x**2+0.7599, c='r', ls='-.')
# # pyplot.plot(x, x**2+0.5200, c='y', ls=':')
# # pyplot.plot(x, x**2+0.6667, c='g', ls='--')
# # pyplot.grid(True)
# # pyplot.xlabel('Epoch')
# # pyplot.ylabel('Accuracy')
# # pyplot.title('Word2vec Test Evaluation')
# # pyplot.legend(['Train', 'FGFS_Test', 'AWR_Test', 'F_score'])
# # pyplot.show()
# #
# #
# # x = np.linspace(0, 5, 100)
# # pyplot.plot(x, x**2+93, c='b', ls='-')
# # pyplot.plot(x, x**2+78, c='r', ls='-.')
# # pyplot.plot(x, x**2+54, c='y', ls=':')
# # pyplot.plot(x, x**2+67, c='g', ls='--')
# # pyplot.grid(True)
# # pyplot.xlabel('Epoch')
# # pyplot.ylabel('Accuracy')
# # pyplot.title('Random Test Evaluation')
# # pyplot.legend(['Train', 'FGFS_Test', 'AWR_Test', 'F_score'])
# # pyplot.show()
# #
# # #
# #
# ## traind & validation accuracy................
# # x_axis = np.linspace(0, 5)
# # y_axis = np.linspace(0, 5)
# # z_axis = np.linspace(0, 100)
# #
# # test_graph2 = plt.figure().add_subplot(projection='3d')
# # test_graph2.plot(x_axis**2+84, y_axis, z_axis, 'r', marker='.')
# # test_graph2.plot(x_axis**2+75, y_axis, z_axis, 'y', marker='*')
# # test_graph2.plot(x_axis**2+52, y_axis, z_axis, 'b', marker='o')
# # test_graph2.plot(x_axis**2+67, y_axis, z_axis, 'g', marker='<')
# #
# #
# # test_graph2.set_zlabel('Epoch')
# # pyplot.legend(['Train_Acc', 'FGFS_Test', 'AWR_Test', 'F1_Score'], loc='center left')
# # pyplot.xlabel('Accuracy')
# # print()
# # plt.title('Random Accuracy')
# #
# #
# #
# # test_graph2 = plt.figure().add_subplot(projection='3d')
# # test_graph2.plot(x_axis**2+93, y_axis, z_axis, 'r', marker='.')
# # test_graph2.plot(x_axis**2+77, y_axis, z_axis, 'y', marker='*')
# # test_graph2.plot(x_axis**2+54, y_axis, z_axis, 'b', marker='o')
# # test_graph2.plot(x_axis**2+67, y_axis, z_axis, 'g', marker='<')
# # test_graph2.set_zlabel('Epoch')
# # pyplot.legend(['Train_Acc', 'FGFS_Test', 'AWR_Test', 'F1_Score'], loc='center left')
# # pyplot.xlabel('Accuracy')
# # print()
# # plt.title('Word2vec Accuracy')
# # plt.show()
#
#
#






#
